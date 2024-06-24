from selenium import webdriver
from __future__ import print_function, unicode_literals
from pprint import pprint
from PyInquirer import prompt, Separator
import bs4
import time
import re
import openpyxl

driver = webdriver.Chrome()

todaysDate = str(r.todaysDate).split()[0]

thisYear = str(r.thisYear)

currency = 'EUR'

# alternative url that will need some math to get the exchange value
# driver.get("https://wise.com/br/pricing/send-money?sourceAmount=100&sourceCcy=EUR&targetCcy=BRL")

driver.get("https://wise.com/tools/exchange-rate-alerts/?fromCurrency=BRL&toCurrency=EUR")

# the link for getting the USD value
# driver.get("https://wise.com/tools/exchange-rate-alerts/?fromCurrency=BRL&toCurrency=USD")

# wait for backend to answer and load infos
time.sleep(3)

if driver.page_source:
  data = bs4.BeautifulSoup(driver.page_source, features="html.parser")
  
  driver.close()
else:
  # TODO: fix this error status
  print(f"Error: {res.status_code }")

wiseScrape = data.select('span.text-success')
# this line returns something like: 
# [<span class="text-success">0.172162 EUR</span>, <span class="text-success">5.80848 BRL</span>]

currencyRegex = r'EUR' if currency == 'EUR' else r'USD'

wiseQuotas = []

wiseQuotas.append(todaysDate)

for i in range(len(wiseScrape)):
  wiseScrape[i] = str(wiseScrape[i])
  # this part saves the currency from the first element on the list (X.XXX EUR/USD)
  tempCurrency = re.search(currencyRegex, wiseScrape[i])
  if (tempCurrency):
    tempCurrency = tempCurrency.group()
    print(tempCurrency)
    wiseQuotas.append(tempCurrency)
  # this part saves the value from the second element of the list (X.XXX BRL)
  tempValue = re.search(r'\d+\.\d+ BRL', wiseScrape[i])
  if (tempValue):
    tempValue = tempValue.group().split()[0]
    print(tempValue)
    wiseQuotas.append(tempValue)

# Adds data to the exchange_historic_series excel spreadsheet

histSeries = openpyxl.load_workbook('exchanges_historic_series.xlsx')

# defines the sheet that should be used based on the year.

try:
  sheet = histSeries[thisYear]
except:
  print(f'Worksheet {thisYear} does not exist')
  print(f'Creating worksheet {thisYear}...')
  
  print('Determining the sheet index...')
  sheetToUse = histSeries.sheetnames
  sheetToUse = sheetToUse[len(sheetToUse) - 1]
  sheet = histSeries[sheetToUse]
  
  histSeries.create_sheet(index=len(sheetToUse)-1, title=thisYear)
  
  sheet = histSeries[thisYear]

sheet.get_highest_column()
sheet['A1'].value
